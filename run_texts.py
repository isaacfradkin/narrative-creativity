"""
Batch pipeline — score every sentence of a stories file with the NT metric.
================================================================================

Reads an .xlsx of participant stories, scores each sentence with the Narrative
Twist metric (see `narrative_twist.py`), and writes three .xlsx files.

Method settings (agreed with Itzik)
-----------------------------------
  - Context S1 = ALL preceding sentences.
  - Continuation S3 = BOTH modes: "next" (next sentence) and "all" (all following).
  - Alternatives S2* : sampling, temperature 1.0, ~10 per sentence.
  - Single score:  nt = log( P(S3|S1,S2) / mean_i P(S3|S1,S2*_i) )   (per-token probs).
  - No aggregation here — one score per sentence (Anna aggregates later in R).

Input
-----
An .xlsx whose first sheet has a `participant` column plus one or more story
columns. Recognized story columns (see KNOWN_STORY_COLS):
  - `paraphrase_text_1`, `paraphrase_text_2`   (the full dataset, 2 stories/participant)
  - `story`                                     (single story per participant)

Outputs (from `--output X.xlsx`)
--------------------------------
  X.xlsx               main : one row per sentence, scores only (nt_next, nt_all).
  X_debug.xlsx         debug: one row per (sentence x S3 mode) + intermediate values.
  X_alternatives.xlsx  validity check: per (sentence x mode), the real sentence and
                       each generated S2*, with its own continuation log-probability.

Example
-------
    python run_texts.py --input "ten stories with clear NTs.xlsx" \
                        --output ten_stories_NT.xlsx --backend llamacpp
"""

import argparse
import re
import time
from typing import List, Tuple

import openpyxl

from narrative_twist import make_backend, narrative_twist


# (short label -> column name in the xlsx). We use whichever are present.
KNOWN_STORY_COLS = [("text_1", "paraphrase_text_1"),
                    ("text_2", "paraphrase_text_2"),
                    ("story", "story")]

# Columns of each output file.
SIMPLE_COLS = ["participant", "story", "sent_idx", "sentence", "nt_next", "nt_all"]
DEBUG_COLS = ["participant", "story", "sent_idx", "sentence", "s3_mode",
              "nt", "log_post", "log_prior", "ratio", "num_alternatives"]
ALT_COLS = ["participant", "story", "sent_idx", "s3_mode", "role", "alt_id",
            "sentence", "log_prob"]


def split_sentences(text) -> List[str]:
    """Split a story into sentences on '.', '?', '!' (followed by whitespace)."""
    if not text:
        return []
    return [p.strip() for p in re.split(r'(?<=[.!?])\s+', str(text).strip()) if p.strip()]


def load_rows(path: str) -> Tuple[List[dict], List[Tuple[str, str]]]:
    """Read the first sheet -> (records, present story columns).

    records: list of {"participant": id, <short label>: story text, ...}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    story_cols = [(short, name) for short, name in KNOWN_STORY_COLS if name in header]
    if not story_cols:
        raise SystemExit(f"No recognized story column in header: {header}")
    col_idx = {name: header.index(name) for _, name in story_cols}
    pid_idx = header.index("participant")
    out = []
    for r in rows[1:]:
        if r[pid_idx] is None:                  # skip blank rows
            continue
        rec = {"participant": r[pid_idx]}
        for short, name in story_cols:
            rec[short] = r[col_idx[name]]
        out.append(rec)
    return out, story_cols


def windows(sentences: List[str]):
    """Yield one sliding window per scorable sentence.

    A target sentence needs at least one sentence before it (context S1) and one
    after it (continuation S3), so t runs from 1 to len-2. For each target we
    yield both continuation definitions:
        (t, S1 = all preceding, S2 = target, S3_next = next, S3_all = all following)
    """
    k = len(sentences)
    for t in range(1, k - 1):
        yield (t,
               " ".join(sentences[:t]),         # S1 = every sentence before t
               sentences[t],                    # S2 = the target sentence
               sentences[t + 1],                # S3 (next) = the next sentence
               " ".join(sentences[t + 1:]))     # S3 (all)  = all following sentences


def write_xlsx(path: str, cols: List[str], rows: List[dict]):
    """Write `rows` (list of dicts) to an .xlsx with the given column order."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    wb.save(path)


def write_simple_xlsx(debug_rows: List[dict], path: str):
    """Pivot the debug rows (one per sentence x mode) into the main file
    (one row per sentence, with nt_next and nt_all side by side)."""
    bucket, order = {}, []
    for r in debug_rows:
        key = (r["participant"], r["story"], r["sent_idx"])
        if key not in bucket:
            bucket[key] = {"participant": r["participant"], "story": r["story"],
                           "sent_idx": r["sent_idx"], "sentence": r["sentence"]}
            order.append(key)
        bucket[key][f"nt_{r['s3_mode']}"] = r["nt"]     # nt_next or nt_all
    write_xlsx(path, SIMPLE_COLS, [bucket[k] for k in order])


def main():
    ap = argparse.ArgumentParser(description="Score every sentence with NT -> 3 xlsx.")
    ap.add_argument("--input", default="Texts.xlsx", help="Input .xlsx of stories.")
    ap.add_argument("--output", default="results.xlsx", help="Base name for the outputs.")
    ap.add_argument("--backend", choices=["hf", "llamacpp"], default="llamacpp",
                    help="hf = distilgpt2 (dev only); llamacpp = Hebrew-Mistral-7B.")
    ap.add_argument("--model", default=None, help="HF model name or GGUF path.")
    ap.add_argument("--n-alternatives", type=int, default=10)
    ap.add_argument("--temperature", type=float, default=1.0)
    ap.add_argument("--seed", type=int, default=42, help="Makes the run reproducible.")
    ap.add_argument("--limit", type=int, default=None, help="Only the first N participants.")
    ap.add_argument("--participants", default=None,
                    help="Comma-separated participant IDs to keep (default: all).")
    args = ap.parse_args()

    rows, story_cols = load_rows(args.input)
    if args.participants:
        keep = {s.strip() for s in args.participants.split(",")}
        rows = [r for r in rows if str(r["participant"]) in keep]
    if args.limit:
        rows = rows[:args.limit]
    print(f"{len(rows)} participant(s) from {args.input}")

    backend = make_backend(args.backend, args.model)

    debug_rows: List[dict] = []                 # one row per (sentence x mode)
    alt_rows: List[dict] = []                   # one row per (sentence x mode x alternative)

    for ri, rec in enumerate(rows, 1):
        pid = rec["participant"]
        for story, _ in story_cols:
            sentences = split_sentences(rec.get(story))
            if len(sentences) < 3:              # need a before and an after -> >=3 sentences
                print(f"  [{ri}/{len(rows)}] {pid}/{story}: {len(sentences)} sentence(s) -> skip",
                      flush=True)
                continue
            print(f"  [{ri}/{len(rows)}] {pid}/{story}: {len(sentences)} sentences", flush=True)
            t_story = time.perf_counter()

            for t, s1, s2, s3_next, s3_all in windows(sentences):
                # Alternatives depend only on S1, so generate them ONCE and reuse
                # them for both S3 modes (saves ~half the 7B generation calls).
                alts = backend.generate_alternatives(
                    s1, args.n_alternatives, temperature=args.temperature, seed=args.seed)
                if not alts:
                    continue

                for mode, s3 in (("next", s3_next), ("all", s3_all)):
                    r = narrative_twist(s1, s2, s3, backend, alternatives=alts, seed=args.seed)
                    if r is None:               # posterior/prior could not be computed
                        continue

                    debug_rows.append({
                        "participant": pid, "story": story, "sent_idx": t,
                        "sentence": s2, "s3_mode": mode,
                        "nt": round(r["nt"], 6),
                        "log_post": round(r["log_post"], 4),
                        "log_prior": round(r["log_prior"], 4),
                        "ratio": round(r["ratio"], 4),
                        "num_alternatives": len(r["prior_details"]),
                    })

                    # Validity-check rows: the real sentence, then each S2*.
                    base = {"participant": pid, "story": story, "sent_idx": t, "s3_mode": mode}
                    alt_rows.append({**base, "role": "real_S2", "alt_id": 0,
                                     "sentence": s2, "log_prob": round(r["log_post"], 4)})
                    for i, (alt_text, lp) in enumerate(r["prior_details"], 1):
                        alt_rows.append({**base, "role": "alternative", "alt_id": i,
                                         "sentence": alt_text, "log_prob": round(lp, 4)})

            print(f"  [{ri}/{len(rows)}] {pid}/{story}: done in "
                  f"{time.perf_counter() - t_story:.1f}s", flush=True)

    base = re.sub(r'\.(xlsx|csv)$', '', args.output)
    write_simple_xlsx(debug_rows, base + ".xlsx")
    write_xlsx(base + "_debug.xlsx", DEBUG_COLS, debug_rows)
    write_xlsx(base + "_alternatives.xlsx", ALT_COLS, alt_rows)
    print(f"\nMain         : {base}.xlsx"
          f"\nDebug        : {base}_debug.xlsx  ({len(debug_rows)} rows)"
          f"\nAlternatives : {base}_alternatives.xlsx  ({len(alt_rows)} rows)")


if __name__ == "__main__":
    main()
